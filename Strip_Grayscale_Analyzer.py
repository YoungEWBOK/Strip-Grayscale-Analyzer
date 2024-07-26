import cv2
import numpy as np
from scipy import signal
from scipy.integrate import simps
import openpyxl
import tifffile
from PIL import Image, ImageDraw, ImageFont
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import os

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

class ImageProcessor:
    def __init__(self, master):
        self.master = master
        self.master.title("条带灰度分析器")
        self.master.geometry("1400x1000")

        self.file_path = None
        self.image = None
        self.processed_image = None
        self.roi = None
        self.hist = None
        self.baseline = 0
        self.orientation = tk.StringVar(value="horizontal")

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=0, column=0, sticky=(tk.N, tk.S))

        ttk.Button(control_frame, text="选择文件", command=self.select_file).pack(pady=5)
        ttk.Button(control_frame, text="处理图像", command=self.process_image).pack(pady=5)

        orientation_frame = ttk.Frame(control_frame)
        orientation_frame.pack(pady=5)
        ttk.Label(orientation_frame, text="框选方向:").pack(side=tk.LEFT)
        ttk.Radiobutton(orientation_frame, text="横向", variable=self.orientation, value="horizontal").pack(side=tk.LEFT)
        ttk.Radiobutton(orientation_frame, text="纵向", variable=self.orientation, value="vertical").pack(side=tk.LEFT)

        ttk.Button(control_frame, text="选择ROI", command=self.select_roi).pack(pady=5)
        ttk.Button(control_frame, text="分析结果", command=self.analyze_results).pack(pady=5)
        ttk.Button(control_frame, text="保存结果", command=self.save_results).pack(pady=5)


        self.image_frame = ttk.Frame(main_frame, width=600, height=400)
        self.image_frame.grid(row=0, column=1, padx=10)

        self.result_frame = ttk.Frame(main_frame, width=600, height=400)
        self.result_frame.grid(row=0, column=2, padx=10)

        self.baseline_frame = ttk.Frame(main_frame, width=1200, height=100)
        self.baseline_frame.grid(row=1, column=1, columnspan=2, pady=10)

        self.results_text_frame = ttk.Frame(main_frame, width=1200, height=100)
        self.results_text_frame.grid(row=2, column=1, columnspan=2, pady=10)

        self.status_var = tk.StringVar()
        ttk.Label(main_frame, textvariable=self.status_var).grid(row=3, column=0, columnspan=3, pady=5)

    def select_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("TIFF files", "*.tif;*.tiff"), ("All files", "*.*")])
        if self.file_path:
            self.status_var.set(f"已选择文件: {self.file_path}")
            self.image = tifffile.imread(self.file_path)
            self.show_image(self.image)

    def process_image(self):
        if self.image is None:
            messagebox.showerror("错误", "请先选择一个文件")
            return

        use_background_subtraction = messagebox.askyesno("背景去除", "是否使用背景去除？\n(Rolling Ball 半径 = 50像素)")
        
        if use_background_subtraction:
            self.processed_image = self.subtract_background(self.image, light_background=True)
        else:
            self.processed_image = self.image

        if len(self.processed_image.shape) > 2:
            self.processed_image = cv2.cvtColor(self.processed_image, cv2.COLOR_RGB2GRAY)

        self.processed_image = self.resize_image(self.processed_image)
        self.show_image(self.processed_image)
        self.status_var.set("图像处理完成")

    def select_roi(self):
        if self.processed_image is None:
            messagebox.showerror("错误", "请先处理图像")
            return

        image_to_show = self.processed_image.copy()
        if len(image_to_show.shape) == 2:
            image_to_show = cv2.cvtColor(image_to_show, cv2.COLOR_GRAY2BGR)

        cv2.namedWindow("Select ROI", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("Select ROI", 800, 600)
        self.roi = cv2.selectROI("Select ROI", image_to_show, False)
        cv2.destroyAllWindows()

        if self.roi == (0, 0, 0, 0):
            self.status_var.set("ROI选择已取消")
            return

        x, y, w, h = self.roi
        roi_image = self.processed_image[y:y+h, x:x+w]
        self.show_image(roi_image)
        self.status_var.set("ROI已选择")

        if self.orientation.get() == "horizontal":
            self.hist = np.sum(roi_image, axis=0)
            self.hist = np.max(self.hist) - self.hist  # Invert the histogram for horizontal selection
        else:  # vertical
            self.hist = np.sum(roi_image, axis=1)
            self.hist = np.max(self.hist) - self.hist  # Invert the histogram for vertical selection

    def analyze_results(self):
        if self.hist is None:
            messagebox.showerror("错误", "请先选择ROI")
            return

        self.plot_results_and_get_baseline()

    def save_results(self):
        if self.hist is None or self.baseline is None:
            messagebox.showerror("错误", "请先分析结果")
            return

        peaks, areas = self.find_peaks_and_calculate_areas(self.hist, self.baseline)
        total_area, max_area, ratio = self.calculate_ratio(areas)

        output_excel = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_excel:
            self.append_to_excel(output_excel, [(len(peaks), total_area, max_area, ratio, areas)])
            self.status_var.set(f"结果已添加到: {output_excel}")

    def show_image(self, image):
        for widget in self.image_frame.winfo_children():
            widget.destroy()

        fig = Figure(figsize=(6, 4))
        ax = fig.add_subplot(111)
        ax.imshow(image, cmap='gray')
        ax.axis('off')

        canvas = FigureCanvasTkAgg(fig, master=self.image_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def resize_image(self, image, max_size=800):
        h, w = image.shape[:2]
        if max(h, w) > max_size:
            if h > w:
                new_h, new_w = max_size, int(max_size * w / h)
            else:
                new_h, new_w = int(max_size * h / w), max_size
            image = cv2.resize(image, (new_w, new_h))
        return image

    def subtract_background(self, image, radius=50, light_background=False):
        # Ensure the image is in 8-bit format
        if image.dtype != np.uint8:
            image = cv2.normalize(image, None, 0, 255, cv2.NORM_MINMAX, dtype=cv2.CV_8U)
        
        # Create a structuring element for morphological operations
        selem = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (radius, radius))
        
        # Estimate the background
        if light_background:
            background = cv2.morphologyEx(image, cv2.MORPH_DILATE, selem)
        else:
            background = cv2.morphologyEx(image, cv2.MORPH_ERODE, selem)
        
        # Calculate the difference between the original image and the background
        diff = cv2.absdiff(image, background)
        
        # Create a mask for the foreground
        _, mask = cv2.threshold(diff, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Dilate the mask to include slightly more of the foreground
        mask = cv2.dilate(mask, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3)), iterations=2)
        
        # Create the result image (white background)
        result = np.full(image.shape, 255, dtype=np.uint8)
        
        # Copy the foreground from the original image to the result
        np.copyto(result, image, where=(mask>0))
        
        return result
    
    def find_peaks_and_calculate_areas(self, hist, baseline):
        above_baseline = hist > baseline
        crossings = np.where(np.diff(above_baseline))[0]
        
        peaks = []
        areas = []
        
        for i in range(0, len(crossings) - 1, 2):
            start = crossings[i]
            end = crossings[i + 1] if i + 1 < len(crossings) else len(hist) - 1
            
            peak_region = hist[start:end+1]
            peak = start + np.argmax(peak_region)
            
            # 使用梯形法则计算面积，模仿ImageJ的面积计算方式
            x = np.arange(start, end+1)
            y = peak_region - baseline
            area = np.trapz(y, x)
            
            peaks.append(peak)
            areas.append(area)
        
        return peaks, areas
    
    def find_significant_peaks(self, hist, baseline, min_area_ratio=0.05):
        peaks, areas = self.find_peaks_and_calculate_areas(hist, baseline)
        total_area = sum(areas)
        significant_peaks = [peak for peak, area in zip(peaks, areas) if area / total_area > min_area_ratio]
        return significant_peaks

    def calculate_ratio(self, areas):
        total_area = sum(areas)
        max_area = max(areas) if areas else 0
        ratio = max_area / total_area if total_area > 0 else 0
        return total_area, max_area, ratio

    def append_to_excel(self, filename, results):
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["峰值数量", "总面积", "最大面积", "比率"]
            for i, header in enumerate(headers, start=1):
                ws.cell(row=i, column=1, value=header)
        else:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

        # 检查是否为空Excel文件
        if ws.max_column == 1 and ws.max_row == 1 and ws['A1'].value is None:
            headers = ["峰值数量", "总面积", "最大面积", "比率"]
            for i, header in enumerate(headers, start=1):
                ws.cell(row=i, column=1, value=header)

        # 找到最后一列
        last_column = ws.max_column
        new_column = last_column + 1

        for result in results:
            peak_count, total_area, max_area, ratio, areas = result
            
            ws.cell(row=1, column=new_column, value=f"分析结果 {new_column-1}")
            ws.cell(row=2, column=new_column, value=peak_count)
            ws.cell(row=3, column=new_column, value=total_area)
            ws.cell(row=4, column=new_column, value=max_area)
            ws.cell(row=5, column=new_column, value=ratio)
            
            # 单独列出各峰面积
            for i, area in enumerate(areas, start=1):
                ws.cell(row=5+i, column=1, value=f"峰{i}")
                ws.cell(row=5+i, column=new_column, value=area)

        wb.save(filename)

    def plot_results_and_get_baseline(self):
        for widget in self.result_frame.winfo_children():
            widget.destroy()

        for widget in self.baseline_frame.winfo_children():
            widget.destroy()

        self.fig, self.ax = plt.subplots(figsize=(6, 4))
        self.line, = self.ax.plot(self.hist, label='直方图')
        self.ax.set_title("选择基线")
        self.ax.set_xlabel("横向位置" if self.orientation.get() == "horizontal" else "纵向位置")
        self.ax.set_ylabel("平均强度")

        self.baseline_line, = self.ax.plot([0, len(self.hist)], [0, 0], 'r--', label='基线')
        self.ax.legend()

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.result_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        slider_frame = ttk.Frame(self.baseline_frame)
        slider_frame.pack(side=tk.LEFT, padx=10)

        ttk.Label(slider_frame, text="基线值:").pack(side=tk.LEFT)
        self.slider = ttk.Scale(slider_frame, from_=0, to=np.max(self.hist), orient=tk.HORIZONTAL, length=300)
        self.slider.pack(side=tk.LEFT)
        self.slider.set(0)

        self.value_label = ttk.Label(slider_frame, text="0")
        self.value_label.pack(side=tk.LEFT, padx=5)

        self.slider.bind("<B1-Motion>", self.update_baseline)
        self.slider.bind("<ButtonRelease-1>", self.update_baseline)

        confirm_button = ttk.Button(self.baseline_frame, text="确认", command=self.confirm_baseline)
        confirm_button.pack(side=tk.LEFT, padx=10)

    def update_baseline(self, event):
        value = self.slider.get()
        self.value_label.config(text=f"{value:.2f}")
        self.baseline_line.set_ydata([value, value])
        self.canvas.draw_idle()

    def confirm_baseline(self):
        self.baseline = float(self.baseline_frame.winfo_children()[0].winfo_children()[2].cget("text"))
        self.display_results()

    def display_results(self):
        for widget in self.result_frame.winfo_children():
            widget.destroy()

        for widget in self.results_text_frame.winfo_children():
            widget.destroy()

        peaks, areas = self.find_peaks_and_calculate_areas(self.hist, self.baseline)
        total_area, max_area, ratio = self.calculate_ratio(areas)

        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(self.hist, label='直方图')
        ax.axhline(y=self.baseline, color='r', linestyle='--', label='基线')
        ax.plot(peaks, self.hist[peaks], "x", label='峰值')
        ax.set_title("分析结果")
        ax.set_xlabel("位置" if self.orientation.get() == "horizontal" else "行位置")
        ax.set_ylabel("平均强度")
        ax.legend()

        canvas = FigureCanvasTkAgg(fig, master=self.result_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # 创建文本框来显示结果
        results_text = tk.Text(self.results_text_frame, height=10, width=80)
        results_text.pack(pady=10)

        results_text.insert(tk.END, f"总面积: {total_area:.2f}\n")
        results_text.insert(tk.END, f"最大面积: {max_area:.2f}\n")
        results_text.insert(tk.END, f"比率: {ratio:.4f}\n\n")
        
        results_text.insert(tk.END, "峰值详情:\n")
        for i, (peak, area) in enumerate(zip(peaks, areas)):
            results_text.insert(tk.END, f"峰值 {i+1}: 位置 = {peak}, 面积 = {area:.2f}\n")

        results_text.config(state=tk.DISABLED)  # 使文本框只读
        
    def add_text_to_image(self, image, text):
        pil_image = Image.fromarray(image)
        draw = ImageDraw.Draw(pil_image)
        font = ImageFont.load_default()
        text_bbox = draw.textbbox((0, 0), text, font=font)
        text_width = text

def main():
    root = tk.Tk()
    app = ImageProcessor(root)
    root.mainloop()

if __name__ == "__main__":
    main()